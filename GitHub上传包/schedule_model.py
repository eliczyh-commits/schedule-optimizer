from __future__ import annotations

import argparse
import math
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

BASE_DIR = Path(__file__).resolve().parent
SEARCH_DIRS = [BASE_DIR, Path.cwd()]
for LOCAL_LIBS in [candidate / name for candidate in SEARCH_DIRS for name in ("libs", ".python-libs2", ".python-libs")]:
    if LOCAL_LIBS.exists():
        sys.path.insert(0, str(LOCAL_LIBS))

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    print("缺少依赖 openpyxl。请先运行：py -m pip install -r requirements.txt", file=sys.stderr)
    raise SystemExit(2)

try:
    import pulp
except ImportError:  # pragma: no cover
    print("缺少依赖 pulp。请先运行：py -m pip install -r requirements.txt", file=sys.stderr)
    raise SystemExit(2)


UNIT_TONS = 100
FOUR_HUNDRED_GRADES = {"HRB400", "HRB400E"}
PRIORITY_GRADES = {"HRB500E", "T63E/E/G", "T63/E/G"}
SIX_HUNDRED_GRADES = {"T63E/E/G", "T63/E/G"}
T63E_SPECS = {12, 14, 16, 18, 20, 22, 25, 28, 32}
T63_SPEC = {10}


class ModelInputError(ValueError):
    pass


@dataclass(frozen=True)
class ResourceRow:
    period: str
    total: int
    foreign: int
    domestic: int
    hrb500e: int
    six_hundred: int
    four_hundred: int


@dataclass(frozen=True)
class ForeignRow:
    period: str
    line: str
    grade: str
    spec: int
    tons: float


@dataclass(frozen=True)
class DemandRow:
    period: str
    grade: str
    spec: int
    tons: int


@dataclass(frozen=True)
class RatioRow:
    spec: int
    lower: float
    upper: float


@dataclass(frozen=True)
class ForecastRow:
    period: str
    grade: str
    spec: int
    tons: float


def norm_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def norm_line(value: Any) -> str:
    text = norm_text(value).replace(" ", "")
    aliases = {"\u68d2\u4e94A": "\u68d2\u4e94A\u7ebf", "\u68d2\u4e94B": "\u68d2\u4e94B\u7ebf"}
    return aliases.get(text, text)

def norm_grade(value: Any) -> str:
    return norm_text(value).replace(" ", "").upper()


def as_float(value: Any, field: str) -> float:
    if value is None or norm_text(value) == "":
        raise ModelInputError(f"{field} 不能为空")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ModelInputError(f"{field} 必须是数字，当前值：{value}") from exc


def as_spec(value: Any, field: str = "规格") -> int:
    number = as_float(value, field)
    if abs(number - round(number)) > 1e-9:
        raise ModelInputError(f"{field} 必须是整数规格，当前值：{value}")
    return int(round(number))


def tons_from_resource(value: Any, field: str) -> int:
    number = as_float(value, field)
    tons = number * 10000 if abs(number) < 1000 else number
    return int(round(tons))


def tons_to_units(tons: float, field: str) -> int:
    units = tons / UNIT_TONS
    if abs(units - round(units)) > 1e-9:
        raise ModelInputError(f"{field} 必须是 {UNIT_TONS} 吨的整数倍，当前吨位：{tons}")
    return int(round(units))


def sheet_rows(ws, required_headers: list[str] | None = None) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_idx = None
    for idx, row in enumerate(rows[:30]):
        names = [norm_text(v) for v in row]
        if required_headers:
            if all(header in names for header in required_headers):
                header_idx = idx
                break
        elif any(names):
            header_idx = idx
            break
    if header_idx is None:
        return []
    headers = [norm_text(v) for v in rows[header_idx]]
    output = []
    for row in rows[header_idx + 1 :]:
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        if any(v is not None and norm_text(v) != "" for v in item.values()):
            output.append(item)
    return output


def find_sheet(wb, names: list[str], required: bool = True):
    name_map = {norm_text(name): name for name in wb.sheetnames}
    for name in names:
        if name in name_map:
            return wb[name_map[name]]
    if required:
        raise ModelInputError(f"缺少工作表：{' 或 '.join(names)}")
    return None


def first_present(row: dict[str, Any], names: list[str], field: str) -> Any:
    for name in names:
        if name in row and row[name] is not None and norm_text(row[name]) != "":
            return row[name]
    raise ModelInputError(f"缺少字段或字段为空：{field}")


def parse_resources(ws) -> list[ResourceRow]:
    rows = sheet_rows(ws)
    if rows and "旬度" in rows[0]:
        parsed = []
        for idx, row in enumerate(rows, start=2):
            period = norm_text(first_present(row, ["旬度"], "旬度"))
            total = tons_from_resource(first_present(row, ["总资源量", "总资源"], "总资源量"), f"旬度资源说明第{idx}行总资源量")
            foreign = tons_from_resource(first_present(row, ["外贸资源量", "外贸"], "外贸资源量"), f"旬度资源说明第{idx}行外贸资源量")
            domestic = tons_from_resource(first_present(row, ["内贸资源量", "内贸"], "内贸资源量"), f"旬度资源说明第{idx}行内贸资源量")
            hrb500e = tons_from_resource(first_present(row, ["500兆帕保供量", "HRB500E保供量", "HRB500E"], "500兆帕保供量"), f"旬度资源说明第{idx}行500兆帕保供量")
            six = tons_from_resource(first_present(row, ["600兆帕保供量", "600兆帕", "600MPA"], "600兆帕保供量"), f"旬度资源说明第{idx}行600兆帕保供量")
            four = row.get("HRB400/HRB400E资源量")
            four_tons = tons_from_resource(four, f"旬度资源说明第{idx}行HRB400/HRB400E资源量") if four is not None and norm_text(four) != "" else domestic - hrb500e - six
            parsed.append(ResourceRow(period, total, foreign, domestic, hrb500e, six, four_tons))
        return parsed

    period = norm_text(ws["C1"].value) or "上旬"
    total = foreign = domestic = None
    hrb500e = six = four = 0
    for row in ws.iter_rows(values_only=True):
        cells = [norm_text(v) for v in row]
        label = "".join(cells[:2])
        values = [v for v in row if isinstance(v, (int, float))]
        if len(cells) > 1 and cells[1] == "总资源" and values:
            total = tons_from_resource(values[-1], "总资源")
        elif "外贸" in label and values:
            foreign = tons_from_resource(values[-1], "外贸资源")
        elif "内贸" in label and values:
            domestic = tons_from_resource(values[-1], "内贸资源")
        elif ("500兆帕" in label or "500MPA" in label.upper() or "HRB500E" in label) and values:
            hrb500e = tons_from_resource(values[-1], "500兆帕保供量")
        elif ("600兆帕" in label or "600MPA" in label.upper() or "T63" in label) and values:
            six += tons_from_resource(values[-1], "600兆帕保供量")
        elif ("HRB400/HRB400E" in label or "HRB400、HRB400E" in label) and len(cells) > 1 and cells[1] == "螺纹" and values:
            four = tons_from_resource(values[-1], "HRB400/HRB400E资源量")
    if total is None or foreign is None or domestic is None:
        raise ModelInputError("旬度资源说明无法识别总资源、外贸资源、内贸资源")
    if four == 0:
        four = domestic - hrb500e - six
    return [ResourceRow(period, total, foreign, domestic, hrb500e, six, four)]


def parse_calendar(ws, default_period: str) -> dict[tuple[str, str], float]:
    rows = sheet_rows(ws, ["产线", "生产天数"])
    result = {}
    for idx, row in enumerate(rows, start=2):
        raw_line = row.get("产线") or row.get("生产条线")
        if raw_line is None or norm_text(raw_line) == "":
            continue
        period = norm_text(row.get("旬度")) or default_period
        line = norm_line(raw_line)
        if row.get("生产天数") is None or norm_text(row.get("生产天数")) == "":
            continue
        days = as_float(first_present(row, ["生产天数", period], "生产天数"), f"生产日历第{idx}行生产天数")
        result[(period, line)] = days
    if result:
        return result

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        result[(default_period, norm_line(row[0]))] = as_float(row[1], "生产天数")
    return result


def parse_line_capacity(ws) -> dict[str, float]:
    rows = sheet_rows(ws, ["产线", "日产能"])
    result = {}
    for idx, row in enumerate(rows, start=2):
        line = norm_line(first_present(row, ["产线", "生产条线"], "产线"))
        result[line] = as_float(first_present(row, ["日产能", "日均产能"], "日产能"), f"产线基础信息第{idx}行日产能")
    return result


def parse_foreign(ws) -> list[ForeignRow]:
    result = []
    for idx, row in enumerate(sheet_rows(ws, ["旬度", "产线", "牌号", "规格", "吨位"]), start=2):
        result.append(
            ForeignRow(
                norm_text(first_present(row, ["旬度"], "旬度")),
                norm_line(first_present(row, ["产线", "生产条线"], "产线")),
                norm_grade(first_present(row, ["牌号"], "牌号")),
                as_spec(first_present(row, ["规格"], "规格")),
                as_float(first_present(row, ["吨位", "外贸吨位"], "吨位"), f"外贸排产明细第{idx}行吨位"),
            )
        )
    return result


def parse_efficiency(ws) -> dict[tuple[str, str, int], float]:
    result = {}
    for idx, row in enumerate(sheet_rows(ws, ["产线", "牌号", "规格", "日产量"]), start=2):
        line = norm_line(first_present(row, ["产线", "生产条线"], "产线"))
        grade = norm_grade(first_present(row, ["牌号"], "牌号"))
        spec = as_spec(first_present(row, ["规格"], "规格"))
        daily = as_float(first_present(row, ["日产量", "日均产量"], "日产量"), f"产品生产效率表第{idx}行日产量")
        if daily <= 0:
            raise ModelInputError(f"产品生产效率表第{idx}行日产量必须大于0")
        result[(line, grade, spec)] = daily
    return result


def parse_cashflow(ws) -> dict[tuple[str, int], float]:
    result = {}
    for idx, row in enumerate(sheet_rows(ws, ["牌号", "规格", "现金流"]), start=2):
        grade = norm_grade(first_present(row, ["牌号"], "牌号"))
        spec = as_spec(first_present(row, ["规格"], "规格"))
        result[(grade, spec)] = as_float(first_present(row, ["现金流"], "现金流"), f"现金流量表第{idx}行现金流")
    return result


def parse_demands(ws) -> list[DemandRow]:
    if ws is None:
        return []
    result = []
    last_period = ""
    last_grade = ""
    for idx, row in enumerate(sheet_rows(ws, ["旬度", "牌号", "规格"]), start=2):
        period = norm_text(row.get("旬度")) or last_period
        grade = norm_grade(row.get("牌号")) or last_grade
        if period:
            last_period = period
        if grade:
            last_grade = grade
        raw_tons = row.get("需求吨位") if "需求吨位" in row else row.get("吨位")
        if raw_tons is None or norm_text(raw_tons) == "":
            continue
        if grade not in PRIORITY_GRADES:
            raise ModelInputError(f"保供量第{idx}行牌号不允许：{grade}")
        spec = as_spec(first_present(row, ["规格"], "规格"))
        if grade == "T63E/E/G" and spec not in T63E_SPECS:
            raise ModelInputError(f"T63E/E/G 只允许 12 至 32 规格，当前规格：{spec}")
        if grade == "T63/E/G" and spec not in T63_SPEC:
            raise ModelInputError(f"T63/E/G 只允许 10 规格，当前规格：{spec}")
        tons = as_float(raw_tons, f"保供量第{idx}行需求吨位")
        result.append(DemandRow(period, grade, spec, tons_to_units(tons, f"保供量第{idx}行需求吨位") * UNIT_TONS))
    return result


def parse_ratios(wb, resource_ws) -> list[RatioRow]:
    ws = find_sheet(wb, ["400兆帕规格比例约束", "规格比例约束"], required=False)
    if ws is not None:
        ratios = []
        for idx, row in enumerate(sheet_rows(ws, ["牌号", "规格", "比例下限", "比例上限"]), start=2):
            grade_group = norm_text(row.get("牌号"))
            if grade_group and "400" not in grade_group and "HRB400" not in grade_group:
                continue
            if row.get("规格") is None or row.get("比例下限") is None or row.get("比例上限") is None:
                continue
            ratios.append(
                RatioRow(
                    as_spec(first_present(row, ["规格"], "规格")),
                    as_float(first_present(row, ["比例下限", "下限"], "比例下限"), f"400兆帕规格比例约束第{idx}行比例下限"),
                    as_float(first_present(row, ["比例上限", "上限"], "比例上限"), f"400兆帕规格比例约束第{idx}行比例上限"),
                )
            )
        if ratios:
            return ratios

    ratios = []
    for row in resource_ws.iter_rows(values_only=True):
        cells = list(row)
        if len(cells) >= 4 and isinstance(cells[1], (int, float)) and isinstance(cells[2], (int, float)) and isinstance(cells[3], (int, float)):
            ratios.append(RatioRow(as_spec(cells[1]), float(cells[2]), float(cells[3])))
    if not ratios:
        raise ModelInputError("缺少 400兆帕规格比例约束，且无法从旬度资源说明读取规格比例")
    return ratios


def parse_forecast(ws, default_period: str) -> list[ForecastRow]:
    if ws is None:
        return []
    result = []
    for idx, row in enumerate(sheet_rows(ws, ["牌号", "规格"]), start=2):
        raw_tons = row.get("客户预报吨位")
        if raw_tons is None:
            raw_tons = row.get("预报吨位")
        if raw_tons is None:
            raw_tons = row.get("需求吨位")
        if raw_tons is None or norm_text(raw_tons) == "":
            continue
        grade = norm_grade(first_present(row, ["牌号"], "牌号"))
        if grade not in FOUR_HUNDRED_GRADES:
            raise ModelInputError(f"400兆帕客户需求预报第{idx}行牌号只能是 HRB400 或 HRB400E")
        result.append(
            ForecastRow(
                norm_text(row.get("旬度")) or default_period,
                grade,
                as_spec(first_present(row, ["规格"], "规格")),
                as_float(raw_tons, f"400兆帕客户需求预报第{idx}行客户预报吨位"),
            )
        )
    return result


def compute_foreign_days(foreign: list[ForeignRow], efficiency: dict[tuple[str, str, int], float]) -> dict[tuple[str, str], float]:
    days = defaultdict(float)
    missing = []
    for row in foreign:
        key = (row.line, row.grade, row.spec)
        if key not in efficiency:
            missing.append(f"产线={row.line}，牌号={row.grade}，规格={row.spec}")
            continue
        days[(row.period, row.line)] += row.tons / efficiency[key]
    if missing:
        unique_missing = sorted(set(missing))
        detail = "；".join(unique_missing[:20])
        if len(unique_missing) > 20:
            detail += f"；另有{len(unique_missing) - 20}项"
        raise ModelInputError(f"外贸排产缺少生产效率：{detail}")
    return dict(days)


def eligible_lines(grade: str, spec: int, efficiency: dict[tuple[str, str, int], float]) -> list[str]:
    return sorted({line for line, g, s in efficiency if g == grade and s == spec})




def ratio_percent_to_fraction(value: float) -> float:
    """Accept either 10 for 10% or 0.1 for 10%."""
    return value / 100 if abs(value) > 1 else value


def format_pct(value: float) -> str:
    number = value * 100
    text = f"{number:.2f}".rstrip("0").rstrip(".")
    return f"{text}%"


def cn(text: str) -> str:
    try:
        return text.encode("ascii").decode("unicode_escape")
    except UnicodeEncodeError:
        return text


def diag_item(kind: str, target: str, status: str, note: str) -> dict[str, Any]:
    return {
        cn("\u5206\u6790\u9879"): kind,
        cn("\u5bf9\u8c61"): target,
        cn("\u5224\u65ad"): status,
        cn("\u8bf4\u660e"): note,
    }




def normalized_grade_ratio_rules(grade_ratios: dict[str, tuple[float, float]] | None) -> dict[str, tuple[float, float]]:
    rules = dict(grade_ratios or {})
    if not rules:
        rules = {"HRB400E": (65.0, 100.0)}
    return rules


def diagnose_relaxed_conflicts(
    resource: ResourceRow,
    calendar: dict[tuple[str, str], float],
    foreign_days: dict[tuple[str, str], float],
    efficiency: dict[tuple[str, str, int], float],
    demands: list[DemandRow],
    ratios: list[RatioRow],
    grade_ratios: dict[str, tuple[float, float]] | None = None,
) -> list[dict[str, Any]]:
    period = resource.period
    period_lines = sorted({line for p, line in calendar if p == period})
    if not period_lines:
        return []

    four_units = tons_to_units(resource.four_hundred, f"{period} 400MPa") if resource.four_hundred else 0
    demand_for_period = [d for d in demands if d.period == period]
    problem = pulp.LpProblem(f"diagnose_{period}", pulp.LpMinimize)
    variables: dict[tuple[str, int, str], pulp.LpVariable] = {}
    slacks: list[tuple[str, str, str, str, pulp.LpVariable, float]] = []

    def add_var(grade: str, spec: int, line: str) -> pulp.LpVariable:
        key = (grade, spec, line)
        if key not in variables:
            name = f"d_{period}_{grade}_{spec}_{line}".replace("/", "_")
            variables[key] = pulp.LpVariable(name, lowBound=0, cat=pulp.LpContinuous)
        return variables[key]

    def add_slack(kind: str, target: str, status: str, note_template: str, weight: float) -> pulp.LpVariable:
        slack = pulp.LpVariable(f"s_{len(slacks)}", lowBound=0, cat=pulp.LpContinuous)
        slacks.append((kind, target, status, note_template, slack, weight))
        return slack

    for demand in demand_for_period:
        lines = eligible_lines(demand.grade, demand.spec, efficiency)
        if not lines:
            continue
        demand_vars = [add_var(demand.grade, demand.spec, line) for line in lines]
        short = add_slack(cn("\u4fdd\u4f9b\u9700\u6c42\u7f3a\u53e3"), f"{demand.grade}-{demand.spec}", cn("\u5173\u952e\u539f\u56e0"), cn("\u8be5\u4fdd\u4f9b\u9700\u6c42\u81f3\u5c11\u8fd8\u7f3a {value} \u5428\u624d\u80fd\u6392\u6ee1\u3002"), 2000)
        over = add_slack(cn("\u4fdd\u4f9b\u9700\u6c42\u8d85\u6392"), f"{demand.grade}-{demand.spec}", cn("\u8f85\u52a9\u5224\u65ad"), cn("\u8be5\u4fdd\u4f9b\u9700\u6c42\u9700\u8981\u5141\u8bb8\u8d85\u6392 {value} \u5428\u624d\u53ef\u5e73\u8861\u3002"), 2000)
        problem += pulp.lpSum(demand_vars) + short / UNIT_TONS - over / UNIT_TONS == tons_to_units(demand.tons, f"{period} {demand.grade} {demand.spec}")

    ratio_specs = sorted({r.spec for r in ratios})
    for grade in sorted(FOUR_HUNDRED_GRADES):
        for spec in ratio_specs:
            for line in eligible_lines(grade, spec, efficiency):
                add_var(grade, spec, line)

    four_vars = [var for (grade, _spec, _line), var in variables.items() if grade in FOUR_HUNDRED_GRADES]
    four_short = add_slack(cn("400\u5146\u5e15\u8d44\u6e90"), period, cn("\u5173\u952e\u539f\u56e0"), cn("400\u5146\u5e15\u8d44\u6e90\u81f3\u5c11\u6709 {value} \u5428\u65e0\u6cd5\u6392\u5165\u73b0\u6709\u4ea7\u7ebf/\u6bd4\u4f8b\u7ec4\u5408\u3002"), 1500)
    four_over = add_slack(cn("400\u5146\u5e15\u8d44\u6e90"), period, cn("\u5173\u952e\u539f\u56e0"), cn("400\u5146\u5e15\u8d44\u6e90\u9700\u8981\u5141\u8bb8\u8d85\u6392 {value} \u5428\u624d\u53ef\u6ee1\u8db3\u5176\u4ed6\u7ea6\u675f\u3002"), 1500)
    problem += pulp.lpSum(four_vars) + four_short / UNIT_TONS - four_over / UNIT_TONS == four_units

    for ratio in ratios:
        lower = ratio_percent_to_fraction(ratio.lower)
        upper = ratio_percent_to_fraction(ratio.upper)
        if not 0 <= lower <= upper <= 1:
            continue
        spec_vars = [var for (grade, spec, _line), var in variables.items() if grade in FOUR_HUNDRED_GRADES and spec == ratio.spec]
        min_units = math.ceil(lower * four_units - 1e-9)
        max_units = math.floor(upper * four_units + 1e-9)
        lower_slack = add_slack(cn("400\u89c4\u683c\u6bd4\u4f8b\u4e0b\u9650"), str(ratio.spec), cn("\u5173\u952e\u539f\u56e0"), cn("\u8be5\u89c4\u683c\u4e0b\u9650\u8981\u6c42\u504f\u9ad8\uff0c\u81f3\u5c11\u7f3a {value} \u5428\u624d\u80fd\u8fbe\u5230\u4e0b\u9650\u3002"), 1000)
        upper_slack = add_slack(cn("400\u89c4\u683c\u6bd4\u4f8b\u4e0a\u9650"), str(ratio.spec), cn("\u5173\u952e\u539f\u56e0"), cn("\u8be5\u89c4\u683c\u4e0a\u9650\u504f\u7d27\uff0c\u81f3\u5c11\u9700\u8981\u653e\u5bbd {value} \u5428\u624d\u6392\u5f97\u4e0b\u3002"), 1000)
        problem += pulp.lpSum(spec_vars) + lower_slack / UNIT_TONS >= min_units
        problem += pulp.lpSum(spec_vars) - upper_slack / UNIT_TONS <= max_units

    for grade, (raw_lower, raw_upper) in normalized_grade_ratio_rules(grade_ratios).items():
        grade = norm_grade(grade)
        if grade not in FOUR_HUNDRED_GRADES:
            continue
        lower = ratio_percent_to_fraction(float(raw_lower))
        upper = ratio_percent_to_fraction(float(raw_upper))
        if not 0 <= lower <= upper <= 1:
            continue
        for spec in ratio_specs:
            same_spec_vars = [var for (var_grade, var_spec, _line), var in variables.items() if var_grade in FOUR_HUNDRED_GRADES and var_spec == spec]
            grade_spec_vars = [var for (var_grade, var_spec, _line), var in variables.items() if var_grade == grade and var_spec == spec]
            if not same_spec_vars:
                continue
            lower_slack = add_slack(cn("400\u724c\u53f7\u6bd4\u4f8b\u4e0b\u9650"), f"{spec}-{grade}", cn("\u5173\u952e\u539f\u56e0"), f"{spec}?? {grade} " + cn("\u4e0b\u9650\u8981\u6c42\u504f\u9ad8\uff0c\u81f3\u5c11\u7f3a {value} \u5428\u624d\u80fd\u8fbe\u5230\u8be5\u89c4\u683c\u5185\u7684\u724c\u53f7\u6bd4\u4f8b\u4e0b\u9650\u3002"), 1000)
            upper_slack = add_slack(cn("400\u724c\u53f7\u6bd4\u4f8b\u4e0a\u9650"), f"{spec}-{grade}", cn("\u5173\u952e\u539f\u56e0"), f"{spec}?? {grade} " + cn("\u4e0a\u9650\u504f\u7d27\uff0c\u81f3\u5c11\u9700\u8981\u653e\u5bbd {value} \u5428\u624d\u6392\u5f97\u4e0b\u3002"), 1000)
            problem += pulp.lpSum(grade_spec_vars) + lower_slack / UNIT_TONS >= lower * pulp.lpSum(same_spec_vars)
            problem += pulp.lpSum(grade_spec_vars) - upper_slack / UNIT_TONS <= upper * pulp.lpSum(same_spec_vars)

    for line in period_lines:
        capacity_days = calendar[(period, line)]
        fixed_days = foreign_days.get((period, line), 0.0)
        line_terms = []
        for (grade, spec, var_line), var in variables.items():
            if var_line == line:
                daily = efficiency.get((line, grade, spec))
                if daily:
                    line_terms.append(var * UNIT_TONS / daily)
        line_slack = add_slack(cn("\u4ea7\u7ebf\u5929\u6570\u4e0d\u8db3"), line, cn("\u5173\u952e\u539f\u56e0"), cn("\u8be5\u4ea7\u7ebf\u81f3\u5c11\u8fd8\u5dee {value} \u5929\uff0c\u8bf4\u660e\u8d44\u6e90/\u4fdd\u4f9b/\u89c4\u683c\u7ec4\u5408\u8d85\u51fa\u8be5\u4ea7\u7ebf\u53ef\u7528\u80fd\u529b\u3002"), 3000)
        problem += pulp.lpSum(line_terms) <= capacity_days - fixed_days + line_slack

    problem += pulp.lpSum(slack * weight for *_meta, slack, weight in slacks)
    status = problem.solve(pulp.PULP_CBC_CMD(msg=False))
    if pulp.LpStatus[status] != "Optimal":
        return [diag_item(cn("\u5173\u952e\u51b2\u7a81\u5b9a\u4f4d"), period, cn("\u65e0\u6cd5\u5206\u6790"), cn("\u8bca\u65ad\u6a21\u578b\u4e5f\u65e0\u6cd5\u6c42\u89e3\uff0c\u8bf7\u4f18\u5148\u68c0\u67e5\u662f\u5426\u7f3a\u5c11\u65e5\u4ea7\u91cf\u3001\u751f\u4ea7\u65e5\u5386\u6216\u9700\u6c42\u8868\u3002"))]

    results: list[dict[str, Any]] = []
    for kind, target, status_text, note_template, slack, _weight in slacks:
        raw = slack.value() or 0
        if raw <= 1e-6:
            continue
        if kind == cn("\u4ea7\u7ebf\u5929\u6570\u4e0d\u8db3"):
            note = note_template.format(value=round(raw, 2))
        else:
            note = note_template.format(value=int(math.ceil(raw - 1e-9)))
        results.append(diag_item(kind, target, status_text, note))
    return results


def diagnose_period_input(
    resource: ResourceRow,
    calendar: dict[tuple[str, str], float],
    foreign_days: dict[tuple[str, str], float],
    efficiency: dict[tuple[str, str, int], float],
    demands: list[DemandRow],
    ratios: list[RatioRow],
    grade_ratios: dict[str, tuple[float, float]] | None = None,
) -> list[dict[str, Any]]:
    period = resource.period
    diagnostics: list[dict[str, Any]] = []
    period_lines = sorted({line for p, line in calendar if p == period})
    available_by_line: dict[str, float] = {}

    if not period_lines:
        diagnostics.append(diag_item(cn("\u751f\u4ea7\u65e5\u5386"), period, cn("\u4e0d\u53ef\u884c"), cn("\u8be5\u65ec\u5ea6\u6ca1\u6709\u586b\u5199\u4efb\u4f55\u4ea7\u7ebf\u751f\u4ea7\u5929\u6570\u3002")))
        return diagnostics

    for line in period_lines:
        capacity_days = calendar[(period, line)]
        occupied = foreign_days.get((period, line), 0.0)
        available = capacity_days - occupied
        available_by_line[line] = max(0.0, available)
        status = cn("\u53ef\u7528") if available >= -1e-8 else cn("\u8d85\u9650")
        diagnostics.append(diag_item(
            cn("\u4ea7\u7ebf\u5269\u4f59\u5929\u6570"),
            line,
            status,
            cn("\u751f\u4ea7\u5929\u6570") + f" {capacity_days:.2f} " + cn("\u5929\uff0c\u5916\u8d38\u5360\u7528") + f" {occupied:.2f} " + cn("\u5929\uff0c\u5185\u8d38\u6700\u591a\u5269\u4f59") + f" {available:.2f} " + cn("\u5929\u3002"),
        ))

    demand_for_period = [d for d in demands if d.period == period]
    for demand in demand_for_period:
        lines = eligible_lines(demand.grade, demand.spec, efficiency)
        if not lines:
            diagnostics.append(diag_item(cn("\u4fdd\u4f9b\u9700\u6c42"), f"{demand.grade}-{demand.spec}", cn("\u4e0d\u53ef\u884c"), cn("\u4ea7\u54c1\u751f\u4ea7\u6548\u7387\u8868\u4e2d\u6ca1\u6709\u4efb\u4f55\u53ef\u6392\u4ea7\u7ebf\uff0c\u65e0\u6cd5\u6392\u5165\u751f\u4ea7\u3002")))
            continue
        best_daily = max(efficiency[(line, demand.grade, demand.spec)] for line in lines)
        best_days = demand.tons / best_daily if best_daily else float("inf")
        eligible_available = sum(available_by_line.get(line, 0.0) for line in lines)
        status = cn("\u9700\u5173\u6ce8") if best_days > eligible_available + 1e-8 else cn("\u53ef\u68c0\u67e5")
        diagnostics.append(diag_item(
            cn("\u4fdd\u4f9b\u9700\u6c42"),
            f"{demand.grade}-{demand.spec}",
            status,
            cn("\u9700\u6c42") + f" {demand.tons} " + cn("\u5428\uff0c\u53ef\u6392\u4ea7\u7ebf\uff1a") + f"{', '.join(lines)}" + cn("\uff1b\u6309\u6700\u5feb\u65e5\u4ea7\u91cf\u81f3\u5c11\u7ea6") + f" {best_days:.2f} " + cn("\u5929\uff0c\u76f8\u5173\u4ea7\u7ebf\u5269\u4f59\u5929\u6570\u5408\u8ba1") + f" {eligible_available:.2f} " + cn("\u5929\u3002"),
        ))

    four_units = tons_to_units(resource.four_hundred, f"{period} 400MPa") if resource.four_hundred else 0
    min_sum = 0
    max_sum = 0
    for ratio in ratios:
        lower = ratio_percent_to_fraction(ratio.lower)
        upper = ratio_percent_to_fraction(ratio.upper)
        min_units = math.ceil(lower * four_units - 1e-9)
        max_units = math.floor(upper * four_units + 1e-9)
        min_sum += min_units
        max_sum += max_units
        spec_lines = sorted({line for grade in FOUR_HUNDRED_GRADES for line in eligible_lines(grade, ratio.spec, efficiency)})
        if lower < -1e-9 or upper > 1 + 1e-9 or lower > upper + 1e-9:
            status = cn("\u4e0d\u53ef\u884c")
            reason = cn("\u6bd4\u4f8b\u4e0a\u4e0b\u9650\u4e0d\u5408\u6cd5\uff0c\u8bf7\u6309\u767e\u5206\u6570\u586b\u5199\uff0c\u4f8b\u5982 10 \u8868\u793a 10%\u3002")
        elif not spec_lines and min_units > 0:
            status = cn("\u4e0d\u53ef\u884c")
            reason = cn("\u8be5\u89c4\u683c\u6709\u6bd4\u4f8b\u4e0b\u9650\uff0c\u4f46\u4ea7\u54c1\u751f\u4ea7\u6548\u7387\u8868\u4e2d\u6ca1\u6709 HRB400/HRB400E \u7684\u53ef\u6392\u4ea7\u7ebf\u3002")
        else:
            status = cn("\u53ef\u68c0\u67e5")
            reason = cn("\u6309 400\u5146\u5e15\u8d44\u6e90") + f" {resource.four_hundred} " + cn("\u5428\u8ba1\u7b97\uff0c\u6700\u4f4e") + f" {min_units * UNIT_TONS} " + cn("\u5428\uff0c\u6700\u9ad8") + f" {max_units * UNIT_TONS} " + cn("\u5428\uff1b\u53ef\u6392\u4ea7\u7ebf\uff1a") + (', '.join(spec_lines) if spec_lines else cn("\u65e0")) + cn("\u3002")
        diagnostics.append(diag_item(cn("400\u89c4\u683c\u6bd4\u4f8b"), str(ratio.spec), status, cn("\u4e0b\u9650") + f" {format_pct(lower)}" + cn("\uff0c\u4e0a\u9650") + f" {format_pct(upper)}" + cn("\u3002") + reason))

    if ratios and min_sum > four_units:
        diagnostics.append(diag_item(cn("400\u89c4\u683c\u6bd4\u4f8b\u6c47\u603b"), period, cn("\u4e0d\u53ef\u884c"), cn("\u6240\u6709\u89c4\u683c\u6700\u4f4e\u5428\u4f4d\u5408\u8ba1") + f" {min_sum * UNIT_TONS} " + cn("\u5428\uff0c\u8d85\u8fc7 400\u5146\u5e15\u8d44\u6e90") + f" {resource.four_hundred} " + cn("\u5428\u3002")))
    if ratios and max_sum < four_units:
        diagnostics.append(diag_item(cn("400\u89c4\u683c\u6bd4\u4f8b\u6c47\u603b"), period, cn("\u4e0d\u53ef\u884c"), cn("\u6240\u6709\u89c4\u683c\u6700\u9ad8\u5428\u4f4d\u5408\u8ba1") + f" {max_sum * UNIT_TONS} " + cn("\u5428\uff0c\u5c0f\u4e8e 400\u5146\u5e15\u8d44\u6e90") + f" {resource.four_hundred} " + cn("\u5428\uff0c\u8d44\u6e90\u65e0\u6cd5\u6392\u6ee1\u3002")))

    e_lines = sorted({line for (line, grade, _spec) in efficiency if grade == "HRB400E"})
    if resource.four_hundred > 0 and not e_lines:
        diagnostics.append(diag_item(cn("HRB400E\u6297\u9707\u5360\u6bd4"), period, cn("\u4e0d\u53ef\u884c"), cn("400\u5146\u5e15\u8981\u6c42 HRB400E \u5360\u6bd4\u81f3\u5c11 65%\uff0c\u4f46\u4ea7\u54c1\u751f\u4ea7\u6548\u7387\u8868\u6ca1\u6709\u4efb\u4f55 HRB400E \u53ef\u6392\u4ea7\u8bb0\u5f55\u3002")))
    elif resource.four_hundred > 0:
        diagnostics.append(diag_item(cn("HRB400E\u6297\u9707\u5360\u6bd4"), period, cn("\u53ef\u68c0\u67e5"), cn("\u81f3\u5c11\u9700\u8981 HRB400E") + f" {math.ceil(0.65 * four_units - 1e-9) * UNIT_TONS} " + cn("\u5428\uff1b\u53ef\u6392 HRB400E \u7684\u4ea7\u7ebf\uff1a") + f"{', '.join(e_lines)}" + cn("\u3002")))

    if not any(row.get(cn("\u5224\u65ad")) == cn("\u4e0d\u53ef\u884c") for row in diagnostics):
        diagnostics.append(diag_item(cn("\u7efc\u5408\u5224\u65ad"), period, cn("\u9700\u7efc\u5408\u8c03\u6574"), cn("\u5355\u9879\u68c0\u67e5\u672a\u53d1\u73b0\u7edd\u5bf9\u9519\u8bef\uff0c\u901a\u5e38\u662f\u591a\u4e2a\u7ea6\u675f\u53e0\u52a0\u5bfc\u81f4\uff1a\u4fdd\u4f9b\u5360\u7528\u4ea7\u7ebf\u540e\uff0c400\u5146\u5e15\u6bd4\u4f8b\u3001HRB400E 65%\u5360\u6bd4\u4e0e\u5404\u4ea7\u7ebf\u65e5\u4ea7\u91cf\u7ec4\u5408\u65e0\u6cd5\u540c\u65f6\u6ee1\u8db3\u3002\u53ef\u4f18\u5148\u653e\u5bbd\u89c4\u683c\u6bd4\u4f8b\uff0c\u6216\u589e\u52a0\u76f8\u5173\u89c4\u683c/\u724c\u53f7\u7684\u53ef\u6392\u4ea7\u7ebf\u548c\u751f\u4ea7\u5929\u6570\u3002")))
    conflict_rows = diagnose_relaxed_conflicts(resource, calendar, foreign_days, efficiency, demands, ratios, grade_ratios)
    if conflict_rows:
        diagnostics = conflict_rows + diagnostics
    return diagnostics

def solve_period(
    resource: ResourceRow,
    calendar: dict[tuple[str, str], float],
    foreign_days: dict[tuple[str, str], float],
    efficiency: dict[tuple[str, str, int], float],
    cashflow: dict[tuple[str, int], float],
    demands: list[DemandRow],
    ratios: list[RatioRow],
    grade_ratios: dict[str, tuple[float, float]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    period = resource.period
    period_lines = sorted({line for p, line in calendar if p == period})
    if not period_lines:
        raise ModelInputError(f"生产日历中没有旬度 {period} 的产线天数")

    demand_for_period = [d for d in demands if d.period == period]
    if (resource.hrb500e > 0 or resource.six_hundred > 0) and not demand_for_period:
        raise ModelInputError(f"{period} 缺少 HRB500E及600兆帕需求表 数据")

    hrb500e_sum = sum(d.tons for d in demand_for_period if d.grade == "HRB500E")
    six_sum = sum(d.tons for d in demand_for_period if d.grade in SIX_HUNDRED_GRADES)
    if hrb500e_sum != resource.hrb500e:
        raise ModelInputError(f"{period} HRB500E需求合计 {hrb500e_sum} 吨，与保供量 {resource.hrb500e} 吨不一致")
    if six_sum != resource.six_hundred:
        raise ModelInputError(f"{period} 600兆帕需求合计 {six_sum} 吨，与保供量 {resource.six_hundred} 吨不一致")

    total_domestic = resource.hrb500e + resource.six_hundred + resource.four_hundred
    if total_domestic != resource.domestic:
        raise ModelInputError(f"{period} 内贸分项合计 {total_domestic} 吨，与内贸资源量 {resource.domestic} 吨不一致")

    four_units = tons_to_units(resource.four_hundred, f"{period} 400兆帕资源量")
    model = pulp.LpProblem(f"domestic_schedule_{period}", pulp.LpMaximize)
    variables: dict[tuple[str, int, str], pulp.LpVariable] = {}

    def add_var(grade: str, spec: int, line: str) -> pulp.LpVariable:
        name = f"x_{period}_{grade}_{spec}_{line}".replace("/", "_")
        var = pulp.LpVariable(name, lowBound=0, cat=pulp.LpInteger)
        variables[(grade, spec, line)] = var
        return var

    for demand in demand_for_period:
        lines = eligible_lines(demand.grade, demand.spec, efficiency)
        if not lines:
            raise ModelInputError(f"{period} 需求无可排产线：牌号={demand.grade}，规格={demand.spec}")
        demand_vars = [add_var(demand.grade, demand.spec, line) for line in lines]
        model += pulp.lpSum(demand_vars) == tons_to_units(demand.tons, f"{period} {demand.grade} {demand.spec}需求吨位")

    ratio_specs = sorted({r.spec for r in ratios})
    for grade in sorted(FOUR_HUNDRED_GRADES):
        for spec in ratio_specs:
            if (grade, spec) not in cashflow:
                raise ModelInputError(f"现金流量表缺少：牌号={grade}，规格={spec}")
            lines = eligible_lines(grade, spec, efficiency)
            if not lines:
                continue
            for line in lines:
                add_var(grade, spec, line)

    four_vars = [var for (grade, _spec, _line), var in variables.items() if grade in FOUR_HUNDRED_GRADES]
    model += pulp.lpSum(four_vars) == four_units

    for ratio in ratios:
        lower = ratio_percent_to_fraction(ratio.lower)
        upper = ratio_percent_to_fraction(ratio.upper)
        if not 0 <= lower <= upper <= 1:
            raise ModelInputError(f"\u89c4\u683c {ratio.spec} \u7684\u6bd4\u4f8b\u4e0a\u4e0b\u9650\u4e0d\u5408\u6cd5\uff1a\u8bf7\u6309\u767e\u5206\u6570\u586b\u5199\uff0c\u4f8b\u5982 10 \u8868\u793a 10%")
        spec_vars = [var for (grade, spec, _line), var in variables.items() if grade in FOUR_HUNDRED_GRADES and spec == ratio.spec]
        min_units = math.ceil(lower * four_units - 1e-9)
        max_units = math.floor(upper * four_units + 1e-9)
        if not spec_vars and min_units > 0:
            raise ModelInputError(f"400???? {ratio.spec} ?????????????")
        model += pulp.lpSum(spec_vars) >= min_units
        model += pulp.lpSum(spec_vars) <= max_units

    grade_ratio_rules = normalized_grade_ratio_rules(grade_ratios)
    for grade, (raw_lower, raw_upper) in grade_ratio_rules.items():
        grade = norm_grade(grade)
        if grade not in FOUR_HUNDRED_GRADES:
            raise ModelInputError(f"400????????? HRB400 ? HRB400E?{grade}")
        lower = ratio_percent_to_fraction(float(raw_lower))
        upper = ratio_percent_to_fraction(float(raw_upper))
        if not 0 <= lower <= upper <= 1:
            raise ModelInputError(f"{grade} ????????????????????? 65 ?? 65%")
        for spec in ratio_specs:
            same_spec_vars = [var for (var_grade, var_spec, _line), var in variables.items() if var_grade in FOUR_HUNDRED_GRADES and var_spec == spec]
            grade_spec_vars = [var for (var_grade, var_spec, _line), var in variables.items() if var_grade == grade and var_spec == spec]
            if not same_spec_vars:
                continue
            model += pulp.lpSum(grade_spec_vars) >= lower * pulp.lpSum(same_spec_vars)
            model += pulp.lpSum(grade_spec_vars) <= upper * pulp.lpSum(same_spec_vars)

    for line in period_lines:
        capacity_days = calendar[(period, line)]
        fixed_days = foreign_days.get((period, line), 0.0)
        if fixed_days - capacity_days > 1e-8:
            raise ModelInputError(f"{period} {line} 外贸已占用 {fixed_days:.2f} 天，超过生产天数 {capacity_days:.2f} 天")
        line_terms = []
        for (grade, spec, var_line), var in variables.items():
            if var_line != line:
                continue
            daily = efficiency[(line, grade, spec)]
            line_terms.append(var * UNIT_TONS / daily)
        model += pulp.lpSum(line_terms) <= capacity_days - fixed_days

    spec_line_flags: dict[tuple[int, str], pulp.LpVariable] = {}

    cash_objective = pulp.lpSum(
        var * UNIT_TONS * cashflow.get((grade, spec), 0.0)
        for (grade, spec, _line), var in variables.items()
        if grade in FOUR_HUNDRED_GRADES
    )
    model += cash_objective

    status = model.solve(pulp.PULP_CBC_CMD(msg=False))
    if pulp.LpStatus[status] == "Optimal":
        optimal_cash = pulp.value(cash_objective) or 0.0
        model += cash_objective >= optimal_cash - 1e-6
        total_units = max(1, tons_to_units(total_domestic, f"{period} ??????"))
        for spec in sorted({spec for (_grade, spec, _line) in variables}):
            for line in period_lines:
                spec_line_vars = [var for (_grade, var_spec, var_line), var in variables.items() if var_spec == spec and var_line == line]
                if not spec_line_vars:
                    continue
                flag = pulp.LpVariable(f"use_{period}_{spec}_{line}".replace("/", "_"), lowBound=0, upBound=1, cat=pulp.LpBinary)
                spec_line_flags[(spec, line)] = flag
                model += pulp.lpSum(spec_line_vars) <= total_units * flag
        if spec_line_flags:
            model.setObjective(pulp.lpSum(spec_line_flags.values()))
            status = model.solve(pulp.PULP_CBC_CMD(msg=False))
    if pulp.LpStatus[status] != "Optimal":
        diagnostics = diagnose_period_input(resource, calendar, foreign_days, efficiency, demands, ratios, grade_ratios)
        key_kind = cn("\u5206\u6790\u9879")
        key_target = cn("\u5bf9\u8c61")
        key_status = cn("\u5224\u65ad")
        key_note = cn("\u8bf4\u660e")
        details = "\n".join(
            f"- {item[key_kind]}?{item[key_target]}?{item[key_status]}?{item[key_note]}"
            for item in diagnostics
        )
        exc = ModelInputError(f"{period} \u65e0\u53ef\u884c\u6392\u4ea7\u65b9\u6848\uff1a{pulp.LpStatus[status]}\u3002\u8bf7\u67e5\u770b\u4e0b\u65b9\u4e0d\u53ef\u884c\u539f\u56e0\u5206\u6790\u3002\n{details}")
        exc.diagnostics = diagnostics
        raise exc

    plan = []
    for (grade, spec, line), var in sorted(variables.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])):
        value = int(round(var.value() or 0))
        if value > 0:
            plan.append({"旬度": period, "牌号": grade, "规格": spec, "吨位": value * UNIT_TONS, "生产条线": line})

    checks = []
    used_days = defaultdict(float)
    for row in plan:
        used_days[row["生产条线"]] += row["吨位"] / efficiency[(row["生产条线"], row["牌号"], row["规格"])]
    for line in period_lines:
        foreign = foreign_days.get((period, line), 0.0)
        domestic = used_days.get(line, 0.0)
        total = foreign + domestic
        capacity = calendar[(period, line)]
        checks.append({"校验项": "产线生产天数", "旬度": period, "产线": line, "结果": "通过" if total <= capacity + 1e-7 else "超限", "数值": round(total, 4), "说明": f"可用{capacity:.4f}天，外贸{foreign:.4f}天，内贸{domestic:.4f}天"})

    by_grade = defaultdict(int)
    by_spec = defaultdict(int)
    total_cash = 0.0
    for row in plan:
        by_grade[row["牌号"]] += row["吨位"]
        if row["牌号"] in FOUR_HUNDRED_GRADES:
            by_spec[row["规格"]] += row["吨位"]
            total_cash += row["吨位"] * cashflow[(row["牌号"], row["规格"])]

    checks.extend(
        [
            {"校验项": "HRB500E需求", "旬度": period, "产线": "", "结果": "通过", "数值": by_grade["HRB500E"], "说明": f"保供量{resource.hrb500e}吨"},
            {"校验项": "600兆帕需求", "旬度": period, "产线": "", "结果": "通过", "数值": by_grade["T63E/E/G"] + by_grade["T63/E/G"], "说明": f"保供量{resource.six_hundred}吨"},
            {"校验项": "400兆帕资源", "旬度": period, "产线": "", "结果": "通过", "数值": by_grade["HRB400"] + by_grade["HRB400E"], "说明": f"资源量{resource.four_hundred}吨"},
            {"校验项": "400兆帕总现金流", "旬度": period, "产线": "", "结果": "通过", "数值": round(total_cash, 4), "说明": ""},
        ]
    )
    display_grade_rules = normalized_grade_ratio_rules(grade_ratios)
    by_spec_grade = defaultdict(int)
    for row in plan:
        if row[cn("\u724c\u53f7")] in FOUR_HUNDRED_GRADES:
            by_spec_grade[(row[cn("\u89c4\u683c")], row[cn("\u724c\u53f7")])] += row[cn("\u5428\u4f4d")]
    for ratio in ratios:
        spec_total = by_spec[ratio.spec]
        for grade, (raw_lower, raw_upper) in display_grade_rules.items():
            grade = norm_grade(grade)
            if grade in FOUR_HUNDRED_GRADES:
                value = by_spec_grade[(ratio.spec, grade)] / spec_total if spec_total else 0
                checks.append({cn("\u6821\u9a8c\u9879"): cn("400\u5146\u5e15\u89c4\u683c") + f"{ratio.spec}-{grade}" + cn("\u724c\u53f7\u6bd4\u4f8b"), cn("\u65ec\u5ea6"): period, cn("\u4ea7\u7ebf"): "", cn("\u7ed3\u679c"): cn("\u901a\u8fc7"), cn("\u6570\u503c"): round(value, 4), cn("\u8bf4\u660e"): cn("\u8be5\u89c4\u683c\u5185\u8303\u56f4") + f"{format_pct(ratio_percent_to_fraction(float(raw_lower)))}-{format_pct(ratio_percent_to_fraction(float(raw_upper)))}"})

    for ratio in ratios:
        value = by_spec[ratio.spec] / resource.four_hundred if resource.four_hundred else 0
        checks.append({"校验项": f"400兆帕规格{ratio.spec}比例", "旬度": period, "产线": "", "结果": "通过", "数值": round(value, 4), "说明": f"范围{ratio.lower}-{ratio.upper}"})
    return plan, checks


def build_suggestions(plan: list[dict[str, Any]], forecasts: list[ForecastRow]) -> list[dict[str, Any]]:
    plan_sum = defaultdict(int)
    for row in plan:
        if row["牌号"] in FOUR_HUNDRED_GRADES:
            plan_sum[(row["旬度"], row["牌号"], row["规格"])] += row["吨位"]
    suggestions = []
    for item in forecasts:
        planned = plan_sum[(item.period, item.grade, item.spec)]
        discount = planned / item.tons if item.tons else None
        if item.tons <= 0:
            reason = "客户预报吨位为0，无法计算折扣"
        elif planned >= item.tons:
            reason = "方案可覆盖客户预报量"
        elif planned == 0:
            reason = "方案未安排该牌号规格，通常受现金流、比例、抗震占比或产线能力约束影响"
        else:
            reason = "方案低于客户预报，建议按折扣承接；原方案保持现金流最优和约束达标"
        suggestions.append(
            {
                "旬度": item.period,
                "牌号": item.grade,
                "规格": item.spec,
                "客户预报吨位": item.tons,
                "排产方案吨位": planned,
                "折扣": round(discount, 4) if discount is not None else "",
                "简要原因": reason,
            }
        )
    return suggestions


def write_output(path: Path, plan: list[dict[str, Any]], checks: list[dict[str, Any]], suggestions: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "内贸排产方案"

    def write_table(sheet, rows: list[dict[str, Any]], headers: list[str]) -> None:
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        for column_cells in sheet.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 45)

    write_table(ws, plan, ["旬度", "牌号", "规格", "吨位", "生产条线"])
    write_table(wb.create_sheet("方案校验"), checks, ["校验项", "旬度", "产线", "结果", "数值", "说明"])
    write_table(wb.create_sheet("400兆帕需求建议"), suggestions, ["旬度", "牌号", "规格", "客户预报吨位", "排产方案吨位", "折扣", "简要原因"])
    wb.save(path)


def run(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path, data_only=True)
    resource_ws = find_sheet(wb, ["旬度资源说明"])
    resources = parse_resources(resource_ws)
    default_period = resources[0].period

    calendar = parse_calendar(find_sheet(wb, ["生产日历"]), default_period)
    _line_capacity = parse_line_capacity(find_sheet(wb, ["产线基础信息"]))
    foreign = parse_foreign(find_sheet(wb, ["外贸排产明细"]))
    efficiency = parse_efficiency(find_sheet(wb, ["产品生产效率表"]))
    cashflow = parse_cashflow(find_sheet(wb, ["现金流量表"]))
    demands = parse_demands(find_sheet(wb, ["HRB500E及600兆帕需求表", "保供量"], required=False))
    ratios = parse_ratios(wb, resource_ws)
    forecasts = parse_forecast(find_sheet(wb, ["400兆帕客户需求预报"], required=False), default_period)
    foreign_days = compute_foreign_days(foreign, efficiency)

    all_plan = []
    all_checks = []
    for resource in resources:
        plan, checks = solve_period(resource, calendar, foreign_days, efficiency, cashflow, demands, ratios)
        all_plan.extend(plan)
        all_checks.extend(checks)

    suggestions = build_suggestions(all_plan, forecasts)
    write_output(output_path, all_plan, all_checks, suggestions)


def main() -> None:
    parser = argparse.ArgumentParser(description="旬度内贸排产优化模型")
    parser.add_argument("--input", required=True, help="输入 Excel 文件路径")
    parser.add_argument("--output", required=True, help="输出 Excel 文件路径")
    args = parser.parse_args()

    try:
        run(Path(args.input), Path(args.output))
    except ModelInputError as exc:
        print(f"输入数据错误：{exc}", file=sys.stderr)
        raise SystemExit(1)


if __name__ == "__main__":
    main()


