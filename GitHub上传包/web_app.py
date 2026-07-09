from __future__ import annotations

import json
import os
import sys
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse

BASE_DIR = Path(__file__).resolve().parent
SEARCH_DIRS = [BASE_DIR, Path.cwd()]
for local_libs in [candidate / name for candidate in SEARCH_DIRS for name in ("libs", ".python-libs2", ".python-libs")]:
    if local_libs.exists():
        sys.path.insert(0, str(local_libs))

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

import schedule_model as model

STATIC_DIR = BASE_DIR / "web"

EMPTY_DEFAULTS = {
    "resources": [{"period": "上旬", "total": "", "foreign": "", "domestic": "", "hrb500e": "", "six_hundred": "", "four_hundred": ""}],
    "calendar": [{"period": "上旬", "line": "", "days": ""}],
    "foreign": [],
    "efficiency": [],
    "cashflow": [],
    "demands": [],
    "ratios": [],
    "grade_ratios": [{"grade": "HRB400E", "lower": 60, "upper": 100}, {"grade": "HRB400", "lower": 30, "upper": 40}],
    "forecast": [],
    "message": "已打开空模板，请直接在网页中输入数据。",
}


def clean_rows(rows):
    return [
        {key: value for key, value in row.items() if value not in ("", None)}
        for row in rows
        if any(value not in ("", None) for value in row.values())
    ]


def to_float(value, default=0.0):
    if value in ("", None):
        return default
    return float(value)


def to_int(value, default=0):
    if value in ("", None):
        return default
    return int(round(float(value)))


def read_defaults() -> dict:
    return json.loads(json.dumps(EMPTY_DEFAULTS, ensure_ascii=False))


def parse_multipart_file(body: bytes, content_type: str) -> bytes:
    marker = "boundary="
    if marker not in content_type:
        raise model.ModelInputError("导入失败：没有找到上传文件边界。")
    boundary = content_type.split(marker, 1)[1].split(";", 1)[0].strip().strip('"')
    delimiter = ("--" + boundary).encode("utf-8")
    for part in body.split(delimiter):
        if b"filename=" not in part:
            continue
        header_end = part.find(b"\r\n\r\n")
        if header_end < 0:
            continue
        content = part[header_end + 4 :]
        if content.endswith(b"\r\n"):
            content = content[:-2]
        if content.endswith(b"--"):
            content = content[:-2]
        if not content:
            raise model.ModelInputError("导入失败：上传文件为空。")
        return content
    raise model.ModelInputError("导入失败：没有读取到 Excel 文件。")


def row_value(row: dict, names: list[str]):
    for name in names:
        value = row.get(name)
        if value not in ("", None):
            return value
    return ""



def find_rows_in_workbook(wb, preferred_sheets: list[str], required_headers: list[str]) -> list[dict]:
    candidates = []
    for name in preferred_sheets:
        if name in wb.sheetnames:
            candidates.append(wb[name])
    candidates.extend(ws for ws in wb.worksheets if ws not in candidates)
    for ws in candidates:
        rows = model.sheet_rows(ws, required_headers)
        if rows:
            return rows
    return []


def zh(text: str) -> str:
    try:
        return text.encode("ascii").decode("unicode_escape")
    except UnicodeEncodeError:
        return text


def import_excel_rows(table: str, file_bytes: bytes) -> list[dict]:
    if load_workbook is None:
        raise model.ModelInputError(zh("\u5bfc\u5165\u5931\u8d25\uff1a\u672a\u5b89\u88c5 openpyxl \u4f9d\u8d56\u3002"))
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    h_period = zh("\u65ec\u5ea6")
    h_line = zh("\u4ea7\u7ebf")
    h_line2 = zh("\u751f\u4ea7\u6761\u7ebf")
    h_grade = zh("\u724c\u53f7")
    h_spec = zh("\u89c4\u683c")
    h_tons = zh("\u5428\u4f4d")
    h_demand_tons = zh("\u9700\u6c42\u5428\u4f4d")
    h_daily = zh("\u65e5\u4ea7\u91cf")
    h_daily2 = zh("\u65e5\u5747\u4ea7\u91cf")
    h_cashflow = zh("\u73b0\u91d1\u6d41")
    h_lower = zh("\u6bd4\u4f8b\u4e0b\u9650")
    h_upper = zh("\u6bd4\u4f8b\u4e0a\u9650")
    h_forecast_tons = zh("\u5ba2\u6237\u9884\u62a5\u5428\u4f4d")
    h_forecast2 = zh("\u9884\u62a5\u5428\u4f4d")

    if table == "foreign":
        rows = find_rows_in_workbook(wb, [zh("\u5916\u8d38\u6392\u4ea7\u660e\u7ec6")], [h_period, h_line, h_grade, h_spec, h_tons])
        result = [
            {
                "period": model.norm_text(row_value(row, [h_period])),
                "line": model.norm_text(row_value(row, [h_line, h_line2])),
                "grade": model.norm_text(row_value(row, [h_grade])),
                "spec": row_value(row, [h_spec]),
                "tons": row_value(row, [h_tons, h_demand_tons]),
            }
            for row in rows
        ]
    elif table == "efficiency":
        rows = find_rows_in_workbook(wb, [zh("\u4ea7\u54c1\u751f\u4ea7\u6548\u7387\u8868")], [h_line, h_grade, h_spec, h_daily])
        result = [
            {
                "line": model.norm_text(row_value(row, [h_line, h_line2])),
                "grade": model.norm_text(row_value(row, [h_grade])),
                "spec": row_value(row, [h_spec]),
                "daily": row_value(row, [h_daily, h_daily2]),
            }
            for row in rows
        ]
    elif table == "cashflow":
        rows = find_rows_in_workbook(wb, [zh("\u73b0\u91d1\u6d41\u91cf\u8868")], [h_grade, h_spec, h_cashflow])
        result = [
            {
                "grade": model.norm_text(row_value(row, [h_grade])),
                "spec": row_value(row, [h_spec]),
                "cashflow": row_value(row, [h_cashflow]),
            }
            for row in rows
        ]
    elif table == "demands":
        rows = find_rows_in_workbook(wb, [zh("\u4fdd\u4f9b\u91cf"), "HRB500E" + zh("\u53ca600\u5146\u5e15\u9700\u6c42\u8868")], [h_grade, h_spec])
        result = []
        last_period = zh("\u4e0a\u65ec")
        last_grade = ""
        for row in rows:
            period = model.norm_text(row_value(row, [h_period])) or last_period
            grade = model.norm_text(row_value(row, [h_grade])) or last_grade
            if period:
                last_period = period
            if grade:
                last_grade = grade
            tons = row_value(row, [h_demand_tons, h_tons])
            if tons in ("", None):
                continue
            try:
                if float(tons) == 0:
                    continue
            except (TypeError, ValueError):
                pass
            result.append({"period": period, "grade": grade, "spec": row_value(row, [h_spec]), "tons": tons})
    elif table == "ratios":
        rows = find_rows_in_workbook(wb, [zh("\u89c4\u683c\u6bd4\u4f8b\u7ea6\u675f"), "400" + zh("\u5146\u5e15\u89c4\u683c\u6bd4\u4f8b\u7ea6\u675f")], [h_spec, h_lower, h_upper])
        result = []
        for row in rows:
            spec = row_value(row, [h_spec])
            lower = row_value(row, [h_lower])
            upper = row_value(row, [h_upper])
            if spec in ("", None) or lower in ("", None) or upper in ("", None):
                continue
            try:
                if abs(float(spec) - round(float(spec))) > 1e-9:
                    continue
            except (TypeError, ValueError):
                continue
            result.append({"spec": spec, "lower": lower, "upper": upper})
    elif table == "forecast":
        rows = find_rows_in_workbook(wb, ["400" + zh("\u5146\u5e15\u5ba2\u6237\u9700\u6c42\u9884\u62a5")], [h_grade, h_spec])
        result = [
            {
                "period": model.norm_text(row_value(row, [h_period])) or zh("\u4e0a\u65ec"),
                "grade": model.norm_text(row_value(row, [h_grade])),
                "spec": row_value(row, [h_spec]),
                "tons": row_value(row, [h_forecast_tons, h_forecast2, h_demand_tons, h_tons]),
            }
            for row in rows
            if row_value(row, [h_forecast_tons, h_forecast2, h_demand_tons, h_tons]) not in ("", None)
        ]
    else:
        raise model.ModelInputError(zh("\u5bfc\u5165\u5931\u8d25\uff1a\u672a\u77e5\u7684\u8868\u683c\u7c7b\u578b\u3002"))

    result = [row for row in result if any(value not in ("", None) for value in row.values())]
    if not result:
        raise model.ModelInputError(zh("\u5bfc\u5165\u5931\u8d25\uff1a\u6ca1\u6709\u8bc6\u522b\u5230\u6709\u6548\u6570\u636e\uff0c\u8bf7\u68c0\u67e5\u5de5\u4f5c\u8868\u540d\u79f0\u548c\u8868\u5934\u3002"))
    return result



def normalize_plan_rows(rows: list[dict]) -> list[dict]:
    k_period = zh("\u65ec\u5ea6")
    k_grade = zh("\u724c\u53f7")
    k_spec = zh("\u89c4\u683c")
    k_tons = zh("\u5428\u4f4d")
    k_line = zh("\u751f\u4ea7\u6761\u7ebf")
    normalized = []
    for row in rows:
        values = list(row.values())
        normalized.append({
            k_period: row.get(k_period, values[0] if len(values) > 0 else ""),
            k_grade: row.get(k_grade, values[1] if len(values) > 1 else ""),
            k_spec: row.get(k_spec, values[2] if len(values) > 2 else ""),
            k_tons: row.get(k_tons, values[3] if len(values) > 3 else ""),
            k_line: row.get(k_line, values[4] if len(values) > 4 else ""),
        })
    return normalized


def cashflow_category(grade: str) -> str:
    grade = model.norm_grade(grade)
    if grade in model.FOUR_HUNDRED_GRADES:
        return zh("\u0034\u0030\u0030\u5146\u5e15")
    if grade == "HRB500E":
        return "HRB500E"
    if grade in model.SIX_HUNDRED_GRADES:
        return "T63E"
    return grade or zh("\u5176\u4ed6")


def build_cashflow_summary(plan: list[dict], cashflow: dict[tuple[str, int], float]) -> list[dict]:
    cat_400 = zh("\u0034\u0030\u0030\u5146\u5e15")
    cat_other = zh("\u5176\u4ed6")
    key_category = zh("\u5206\u7c7b")
    key_tons = zh("\u5428\u4f4d")
    key_cash = zh("\u73b0\u91d1\u6d41")
    key_note = zh("\u8bf4\u660e")
    key_grade = zh("\u724c\u53f7")
    key_spec = zh("\u89c4\u683c")
    buckets = {
        cat_400: {"tons": 0.0, "cash": 0.0, "missing": set()},
        "HRB500E": {"tons": 0.0, "cash": 0.0, "missing": set()},
        "T63E": {"tons": 0.0, "cash": 0.0, "missing": set()},
    }
    total_tons = 0.0
    total_cash = 0.0
    total_missing = set()
    for row in plan:
        grade = model.norm_grade(row.get(key_grade))
        spec = to_int(row.get(key_spec))
        tons = to_float(row.get(key_tons))
        category = cashflow_category(grade)
        if category not in buckets:
            buckets[category] = {"tons": 0.0, "cash": 0.0, "missing": set()}
        unit_cash = cashflow.get((grade, spec))
        buckets[category]["tons"] += tons
        total_tons += tons
        if unit_cash is None:
            buckets[category]["missing"].add(f"{grade}-{spec}")
            total_missing.add(f"{grade}-{spec}")
            continue
        amount = tons * unit_cash
        buckets[category]["cash"] += amount
        total_cash += amount
    rows = []
    ordered_categories = [cat_400, "HRB500E", "T63E"] + sorted(k for k in buckets if k not in {cat_400, "HRB500E", "T63E"})
    for category in ordered_categories:
        item = buckets[category]
        note = ""
        if item["missing"]:
            note = zh("\u7f3a\u5c11\u73b0\u91d1\u6d41\uff1a") + zh("\u3001").join(sorted(item["missing"]))
        rows.append({key_category: category, key_tons: round(item["tons"], 4), key_cash: round(item["cash"], 4), key_note: note})
    rows.append({
        key_category: zh("\u6c47\u603b"),
        key_tons: round(total_tons, 4),
        key_cash: round(total_cash, 4),
        key_note: "" if not total_missing else zh("\u90e8\u5206\u724c\u53f7\u89c4\u683c\u7f3a\u5c11\u73b0\u91d1\u6d41\uff0c\u6c47\u603b\u672a\u5305\u542b\u7f3a\u5931\u9879"),
    })
    return rows

def workbook_bytes(rows: list[dict], headers: list[str], sheet_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or zh("\u5bfc\u51fa"))[:31]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def solve_payload(payload: dict) -> dict:
    resources = []
    for row in clean_rows(payload.get("resources", [])):
        period = model.norm_text(row.get("period"))
        total = model.tons_from_resource(row.get("total"), f"{period} 总资源量")
        foreign = model.tons_from_resource(row.get("foreign"), f"{period} 外贸资源量")
        domestic = total - foreign
        hrb500e = model.tons_from_resource(row.get("hrb500e"), f"{period} 500兆帕保供量")
        six = model.tons_from_resource(row.get("six_hundred"), f"{period} 600兆帕保供量")
        four_tons = domestic - hrb500e - six
        resources.append(model.ResourceRow(period, total, foreign, domestic, hrb500e, six, four_tons))

    calendar = {}
    for row in clean_rows(payload.get("calendar", [])):
        calendar[(model.norm_text(row.get("period")), model.norm_line(row.get("line")))] = to_float(row.get("days"))

    foreign_rows = []
    for row in clean_rows(payload.get("foreign", [])):
        foreign_rows.append(
            model.ForeignRow(
                model.norm_text(row.get("period")),
                model.norm_line(row.get("line")),
                model.norm_grade(row.get("grade")),
                to_int(row.get("spec")),
                to_float(row.get("tons")),
            )
        )

    efficiency = {}
    for row in clean_rows(payload.get("efficiency", [])):
        efficiency[(model.norm_line(row.get("line")), model.norm_grade(row.get("grade")), to_int(row.get("spec")))] = to_float(row.get("daily"))

    cashflow = {}
    for row in clean_rows(payload.get("cashflow", [])):
        cashflow[(model.norm_grade(row.get("grade")), to_int(row.get("spec")))] = to_float(row.get("cashflow"))

    demands = []
    for row in clean_rows(payload.get("demands", [])):
        tons = to_float(row.get("tons"))
        if tons == 0:
            continue
        demands.append(model.DemandRow(model.norm_text(row.get("period")), model.norm_grade(row.get("grade")), to_int(row.get("spec")), model.tons_to_units(tons, "保供需求吨位") * model.UNIT_TONS))

    ratios = []
    for row in clean_rows(payload.get("ratios", [])):
        ratios.append(model.RatioRow(to_int(row.get("spec")), to_float(row.get("lower")), to_float(row.get("upper"))))

    grade_ratios = {}
    for row in clean_rows(payload.get("grade_ratios", [])):
        grade = model.norm_grade(row.get("grade"))
        if not grade:
            continue
        grade_ratios[grade] = (to_float(row.get("lower")), to_float(row.get("upper")))

    forecasts = []
    for row in clean_rows(payload.get("forecast", [])):
        tons = to_float(row.get("tons"))
        if tons == 0:
            continue
        forecasts.append(model.ForecastRow(model.norm_text(row.get("period")), model.norm_grade(row.get("grade")), to_int(row.get("spec")), tons))

    if not resources:
        raise model.ModelInputError("至少需要填写一行旬度资源")

    foreign_days = model.compute_foreign_days(foreign_rows, efficiency)
    all_plan = []
    all_checks = []
    for resource in resources:
        try:
            plan, checks = model.solve_period(resource, calendar, foreign_days, efficiency, cashflow, demands, ratios, grade_ratios)
        except model.ModelInputError as exc:
            if not hasattr(exc, "diagnostics") and hasattr(model, "diagnose_period_input"):
                exc.diagnostics = model.diagnose_period_input(resource, calendar, foreign_days, efficiency, demands, ratios, grade_ratios)
            raise
        all_plan.extend(plan)
        all_checks.extend(checks)
    all_plan = normalize_plan_rows(all_plan)
    suggestions = model.build_suggestions(all_plan, forecasts)
    cashflow_summary = build_cashflow_summary(all_plan, cashflow)
    return {"plan": all_plan, "checks": all_checks, "suggestions": suggestions, "cashflow_summary": cashflow_summary}


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):  # noqa: A003
        return

    def send_json(self, status: int, body: dict) -> None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_excel(self, filename: str, rows: list[dict], headers: list[str], sheet_name: str) -> None:
        data = workbook_bytes(rows, headers, sheet_name)
        safe_name = quote(filename)
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename*=UTF-8''" + safe_name)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path == "/api/defaults":
            self.send_json(200, read_defaults())
            return
        target = STATIC_DIR / ("index.html" if parsed.path in ("/", "") else parsed.path.lstrip("/"))
        if not target.resolve().is_relative_to(STATIC_DIR.resolve()) or not target.exists():
            self.send_error(404)
            return
        content_type = "text/html; charset=utf-8" if target.suffix == ".html" else "text/plain; charset=utf-8"
        if target.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif target.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        data = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self):  # noqa: N802
        parsed = urlparse(self.path)
        length = int(self.headers.get("Content-Length", "0"))
        try:
            if parsed.path == "/api/solve":
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                self.send_json(200, solve_payload(payload))
                return
            if parsed.path == "/api/export-excel":
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                self.send_excel(payload.get("filename", "export.xlsx"), payload.get("rows", []), payload.get("headers", []), payload.get("sheetName", zh("\u5bfc\u51fa")))
                return
            if parsed.path == "/api/import-excel":
                table = parse_qs(parsed.query).get("table", [""])[0]
                body = self.rfile.read(length)
                file_bytes = parse_multipart_file(body, self.headers.get("Content-Type", ""))
                self.send_json(200, {"rows": import_excel_rows(table, file_bytes)})
                return
            self.send_error(404)
        except model.ModelInputError as exc:
            body = {"error": str(exc)}
            diagnostics = getattr(exc, "diagnostics", None)
            if diagnostics is not None:
                body["diagnostics"] = diagnostics
            self.send_json(400, body)
        except Exception as exc:  # pragma: no cover
            self.send_json(500, {"error": f"程序错误：{exc}"})


def main() -> None:
    port = int(os.environ.get("PORT", "8765"))
    server = ThreadingHTTPServer(("0.0.0.0", port), Handler)
    print(f"排产网页应用已启动：http://127.0.0.1:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()



